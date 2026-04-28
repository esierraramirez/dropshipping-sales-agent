#!/usr/bin/env python3
"""Evaluacion OE4 del agente con 4 KPIs tecnicos (sin satisfaccion)."""

from __future__ import annotations

import argparse
import json
import math
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook


THRESHOLDS = {
    "error_rate_percent_max": 5.0,
    "availability_percent_min": 95.0,
    "latency_p95_seconds_max": 8.0,
    "response_integrity_percent_min": 95.0,
}


@dataclass
class Scenario:
    id: int
    category: str
    title: str
    message: str
    history: list[dict[str, str]] | None = None
    purchase_context: dict[str, Any] | None = None


def build_30_scenarios() -> list[Scenario]:
    scenarios: list[Scenario] = [
        Scenario(1, "precio", "Precio camisa basica", "Hola, cuanto cuesta la camisa basica blanca talla M?"),
        Scenario(2, "precio", "Precio mayorista", "Si compro 10 unidades del set de maquillaje, que precio final manejas?"),
        Scenario(3, "precio", "Precio con promocion", "Tienen algun descuento hoy para las zapatillas urbanas negras?"),
        Scenario(4, "precio", "Precio en rango", "Busco una mochila entre 20 y 30 dolares, que opciones hay?"),
        Scenario(5, "precio", "Precio con variacion", "La chaqueta impermeable cambia de precio por talla o color?"),
        Scenario(6, "disponibilidad", "Stock talla especifica", "Hay stock del vestido floral en talla S para envio inmediato?"),
        Scenario(7, "disponibilidad", "Disponibilidad por color", "Quiero el reloj minimalista en color plata, esta disponible?"),
        Scenario(8, "disponibilidad", "Reposicion", "Cuando vuelve a entrar la cartera premium cafe?"),
        Scenario(9, "disponibilidad", "Ultimas unidades", "Me confirmas si quedan ultimas unidades del audifono bluetooth pro?"),
        Scenario(10, "disponibilidad", "Disponibilidad por cantidad", "Necesito 25 termos personalizados, puedes confirmar disponibilidad?"),
        Scenario(11, "recomendacion", "Recomendacion regalo", "Recomiendame un regalo para mujer de 30 anos, presupuesto 40 dolares."),
        Scenario(12, "recomendacion", "Recomendacion clima", "Que me recomiendas para temporada de lluvia en ciudad humeda?"),
        Scenario(13, "recomendacion", "Recomendacion uso diario", "Busco zapatillas comodas para uso diario, cual conviene?"),
        Scenario(14, "recomendacion", "Recomendacion emprendimiento", "Voy a iniciar dropshipping de accesorios, que 3 productos me sugieres primero?"),
        Scenario(15, "recomendacion", "Recomendacion por perfil", "Soy estudiante y viajo ligero, que mochila me recomiendas?"),
        Scenario(16, "comparacion", "Comparacion dos productos", "Comparame el smartwatch X y smartwatch Y en bateria y precio."),
        Scenario(17, "comparacion", "Comparacion calidad", "Cual es mejor entre la sudadera A y B para clima frio?"),
        Scenario(18, "comparacion", "Comparacion envio", "Entre tenis runner y tenis urban, cual llega mas rapido?"),
        Scenario(19, "comparacion", "Comparacion costo-beneficio", "Quiero la opcion con mejor relacion calidad precio entre estas dos mochilas."),
        Scenario(20, "comparacion", "Comparacion tecnica", "Diferencias clave entre auricular bass y auricular clear en sonido y autonomia?"),
        Scenario(21, "envio", "Condiciones envio local", "Cuanto tarda el envio a Bogota y cual es el costo?"),
        Scenario(22, "envio", "Envio internacional", "Hacen envios internacionales a Peru?"),
        Scenario(23, "envio", "Politica devolucion", "Si el producto llega defectuoso, como funciona el cambio o devolucion?"),
        Scenario(24, "envio", "Seguimiento", "Como puedo rastrear mi pedido una vez pagado?"),
        Scenario(25, "envio", "Envio express", "Hay opcion de envio express para que llegue manana?"),
        Scenario(
            26,
            "compra",
            "Cierre compra simple",
            "Perfecto, quiero comprar 2 unidades del set skincare natural, ayudame a cerrar el pedido.",
            history=[
                {"role": "user", "content": "Me interesa el set skincare natural."},
                {"role": "assistant", "content": "Claro, te comparto precio y detalles del set skincare natural."},
            ],
        ),
        Scenario(
            27,
            "compra",
            "Compra con datos cliente",
            "Confirmo compra de 1 reloj minimalista plata y 1 cartera premium. Procedemos?",
            purchase_context={
                "customer_name": "Cliente Prueba",
                "customer_phone": "+573001112233",
                "customer_address": "Calle 100 #10-20",
                "items": [
                    {
                        "product_id": "R-001",
                        "product_name": "Reloj minimalista plata",
                        "quantity": 1,
                        "unit_price": 29.9,
                    },
                    {
                        "product_id": "C-002",
                        "product_name": "Cartera premium",
                        "quantity": 1,
                        "unit_price": 35.5,
                    },
                ],
                "total_amount": 65.4,
                "is_confirmed": True,
            },
        ),
        Scenario(28, "compra", "Metodo de pago", "Listo, como hago el pago para finalizar mi compra hoy?"),
        Scenario(29, "compra", "Direccion entrega", "Ya elegi el producto, que datos necesitas para generar la orden?"),
        Scenario(30, "compra", "Cierre con urgencia", "Necesito cerrar el pedido en este momento, guiame paso a paso."),
    ]
    if len(scenarios) != 30:
        raise RuntimeError("La bateria de pruebas debe contener exactamente 30 escenarios")
    return scenarios


def percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    if p <= 0:
        return min(values)
    if p >= 100:
        return max(values)
    ordered = sorted(values)
    rank = (len(ordered) - 1) * (p / 100.0)
    lo = math.floor(rank)
    hi = math.ceil(rank)
    if lo == hi:
        return ordered[int(rank)]
    return ordered[lo] + (ordered[hi] - ordered[lo]) * (rank - lo)


def login(client: httpx.Client, base_url: str, email: str, password: str) -> str:
    response = client.post(f"{base_url}/auth/login", json={"email": email, "password": password})
    response.raise_for_status()
    payload = response.json()
    token = payload.get("access_token")
    if not token:
        raise RuntimeError("No se recibio access_token en /auth/login")
    return token


def run_scenarios(base_url: str, email: str, password: str, timeout_seconds: float) -> dict[str, Any]:
    scenarios = build_30_scenarios()
    total_requests = len(scenarios)

    results: list[dict[str, Any]] = []
    errors = 0
    health_checks_ok = 0
    response_integrity_ok = 0
    latencies: list[float] = []

    with httpx.Client(timeout=timeout_seconds) as client:
        token = login(client, base_url, email, password)
        headers = {"Authorization": f"Bearer {token}"}

        for scenario in scenarios:
            health_ok = False
            try:
                health_resp = client.get(f"{base_url}/health")
                health_ok = health_resp.status_code == 200
            except Exception:
                health_ok = False

            if health_ok:
                health_checks_ok += 1

            started = time.perf_counter()
            status_code = None
            error_detail = None
            response_json: dict[str, Any] | None = None

            payload: dict[str, Any] = {"message": scenario.message}
            if scenario.history:
                payload["history"] = scenario.history
            if scenario.purchase_context:
                payload["purchase_context"] = scenario.purchase_context

            try:
                resp = client.post(f"{base_url}/chat/me", headers=headers, json=payload)
                status_code = resp.status_code
                if status_code == 200:
                    response_json = resp.json()
                else:
                    error_detail = resp.text[:500]
            except Exception as exc:
                error_detail = str(exc)

            elapsed = time.perf_counter() - started
            success = status_code == 200 and isinstance(response_json, dict)
            if success:
                latencies.append(elapsed)
            else:
                errors += 1

            agent_response = ""
            context_used = ""
            matches_found = None
            if success:
                agent_response = str(response_json.get("agent_response", "")).strip()
                context_used = str(response_json.get("context_used", "")).strip()
                matches_found = response_json.get("matches_found")

            has_integrity = success and len(agent_response) >= 40 and bool(context_used)
            if has_integrity:
                response_integrity_ok += 1

            results.append(
                {
                    "id": scenario.id,
                    "category": scenario.category,
                    "title": scenario.title,
                    "message": scenario.message,
                    "status_code": status_code,
                    "success": success,
                    "health_ok_before_request": health_ok,
                    "response_time_seconds": round(elapsed, 3),
                    "response_length": len(agent_response),
                    "context_used": context_used,
                    "matches_found": matches_found,
                    "response_integrity_ok": has_integrity,
                    "agent_response": agent_response,
                    "error_detail": error_detail,
                }
            )

    error_rate = (errors / total_requests) * 100.0 if total_requests else 0.0
    availability = (health_checks_ok / total_requests) * 100.0 if total_requests else 0.0
    latency_p95 = percentile(latencies, 95)
    response_integrity = (response_integrity_ok / total_requests) * 100.0 if total_requests else 0.0

    kpis = {
        "error_rate_percent": round(error_rate, 2),
        "availability_percent": round(availability, 2),
        "latency_p95_seconds": round(latency_p95, 3),
        "response_integrity_percent": round(response_integrity, 2),
    }

    kpi_pass = {
        "error_rate": kpis["error_rate_percent"] <= THRESHOLDS["error_rate_percent_max"],
        "availability": kpis["availability_percent"] >= THRESHOLDS["availability_percent_min"],
        "latency_p95": kpis["latency_p95_seconds"] <= THRESHOLDS["latency_p95_seconds_max"],
        "response_integrity": kpis["response_integrity_percent"] >= THRESHOLDS["response_integrity_percent_min"],
    }

    return {
        "metadata": {
            "executed_at": datetime.now().isoformat(timespec="seconds"),
            "base_url": base_url,
            "total_requests": total_requests,
            "thresholds": THRESHOLDS,
        },
        "kpis": kpis,
        "kpi_pass": kpi_pass,
        "results": results,
    }


def render_markdown_report(report: dict[str, Any]) -> str:
    meta = report["metadata"]
    kpis = report["kpis"]
    flags = report["kpi_pass"]

    def pass_label(value: bool) -> str:
        return "PASS" if value else "FAIL"

    lines = [
        "# Reporte OE4 - KPIs tecnicos del agente",
        "",
        f"- Fecha de ejecucion: {meta['executed_at']}",
        f"- Base URL: {meta['base_url']}",
        f"- Total conversaciones: {meta['total_requests']}",
        "",
        "## KPIs",
        "",
        "| KPI | Resultado | Umbral | Estado |",
        "|---|---:|---:|---|",
        f"| Tasa de errores del sistema | {kpis['error_rate_percent']:.2f}% | <= {THRESHOLDS['error_rate_percent_max']:.2f}% | {pass_label(flags['error_rate'])} |",
        f"| Disponibilidad operativa | {kpis['availability_percent']:.2f}% | >= {THRESHOLDS['availability_percent_min']:.2f}% | {pass_label(flags['availability'])} |",
        f"| Latencia p95 del backend | {kpis['latency_p95_seconds']:.3f}s | <= {THRESHOLDS['latency_p95_seconds_max']:.3f}s | {pass_label(flags['latency_p95'])} |",
        f"| Integridad de respuesta del agente | {kpis['response_integrity_percent']:.2f}% | >= {THRESHOLDS['response_integrity_percent_min']:.2f}% | {pass_label(flags['response_integrity'])} |",
    ]
    return "\n".join(lines)


def write_excel_report(report: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_summary.append(["KPI", "Formula", "Resultado", "Umbral", "Cumple"])

    kpis = report["kpis"]
    flags = report["kpi_pass"]
    total = report["metadata"]["total_requests"]

    ws_summary.append([
        "Tasa de errores del sistema",
        "(errores / total solicitudes) * 100",
        kpis["error_rate_percent"],
        f"<= {THRESHOLDS['error_rate_percent_max']}",
        "PASS" if flags["error_rate"] else "FAIL",
    ])
    ws_summary.append([
        "Disponibilidad operativa",
        "(health checks exitosos / total health checks) * 100",
        kpis["availability_percent"],
        f">= {THRESHOLDS['availability_percent_min']}",
        "PASS" if flags["availability"] else "FAIL",
    ])
    ws_summary.append([
        "Latencia p95 del backend",
        "percentil 95 de tiempos de /chat/me",
        kpis["latency_p95_seconds"],
        f"<= {THRESHOLDS['latency_p95_seconds_max']}",
        "PASS" if flags["latency_p95"] else "FAIL",
    ])
    ws_summary.append([
        "Integridad de respuesta del agente",
        "(respuestas con contenido y contexto / total solicitudes) * 100",
        kpis["response_integrity_percent"],
        f">= {THRESHOLDS['response_integrity_percent_min']}",
        "PASS" if flags["response_integrity"] else "FAIL",
    ])
    ws_summary.append(["Total conversaciones", "", total, "", ""])

    ws_cases = wb.create_sheet("Casos")
    ws_cases.append([
        "id",
        "category",
        "title",
        "status_code",
        "success",
        "health_ok_before_request",
        "response_time_seconds",
        "response_length",
        "context_used",
        "matches_found",
        "response_integrity_ok",
        "error_detail",
    ])
    for row in report["results"]:
        ws_cases.append([
            row["id"],
            row["category"],
            row["title"],
            row["status_code"],
            row["success"],
            row["health_ok_before_request"],
            row["response_time_seconds"],
            row["response_length"],
            row["context_used"],
            row["matches_found"],
            row["response_integrity_ok"],
            row["error_detail"],
        ])

    ws_raw = wb.create_sheet("Respuestas")
    ws_raw.append(["id", "message", "agent_response"])
    for row in report["results"]:
        ws_raw.append([row["id"], row["message"], row["agent_response"]])

    wb.save(output_path)


def ensure_output_dir(path: str) -> Path:
    out = Path(path)
    out.mkdir(parents=True, exist_ok=True)
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Ejecuta 30 conversaciones y evalua 4 KPIs tecnicos OE4")
    parser.add_argument("--base-url", default="http://127.0.0.1:8000", help="URL base del backend")
    parser.add_argument("--email", required=True, help="Email del vendor para login")
    parser.add_argument("--password", required=True, help="Password del vendor para login")
    parser.add_argument("--output-dir", default="backed/tests/reports", help="Directorio de reportes")
    parser.add_argument("--timeout", type=float, default=45.0, help="Timeout por request en segundos")
    args = parser.parse_args()

    report = run_scenarios(
        base_url=args.base_url.rstrip("/"),
        email=args.email,
        password=args.password,
        timeout_seconds=args.timeout,
    )

    out_dir = ensure_output_dir(args.output_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = out_dir / f"oe4_eval_kpi4_{stamp}.json"
    md_path = out_dir / f"oe4_eval_kpi4_{stamp}.md"
    xlsx_path = out_dir / f"oe4_eval_kpi4_{stamp}.xlsx"

    json_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    md_path.write_text(render_markdown_report(report), encoding="utf-8")
    write_excel_report(report, xlsx_path)

    kpis = report["kpis"]
    print("=== Evaluacion OE4 completada (4 KPIs tecnicos) ===")
    print(f"Tasa de errores: {kpis['error_rate_percent']:.2f}%")
    print(f"Disponibilidad: {kpis['availability_percent']:.2f}%")
    print(f"Latencia p95: {kpis['latency_p95_seconds']:.3f}s")
    print(f"Integridad de respuesta: {kpis['response_integrity_percent']:.2f}%")
    print(f"Reporte JSON: {json_path}")
    print(f"Reporte MD: {md_path}")
    print(f"Reporte XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
